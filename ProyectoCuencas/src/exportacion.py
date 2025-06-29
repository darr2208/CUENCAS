import io
import geopandas as gpd
import pandas as pd
import tempfile
import zipfile
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference

def exportar_shapefile_zip(gdf):
    with tempfile.TemporaryDirectory() as tmpdir:
        shp_path = os.path.join(tmpdir, "cuenca.shp")
        gdf.to_crs("EPSG:4326").to_file(shp_path, driver="ESRI Shapefile")
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w') as zipf:
            for ext in [".shp", ".shx", ".dbf", ".prj"]:
                file_path = os.path.join(tmpdir, f"cuenca{ext}")
                zipf.write(file_path, arcname=f"cuenca{ext}")
        zip_buffer.seek(0)
        return zip_buffer

def exportar_excel(resultados_dict):
    parametros_geom = {k: v for k, v in resultados_dict.items() if "h)" not in k}
    tiempos_conc = {k: v for k, v in resultados_dict.items() if "h)" in k}

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Parámetros Geométricos"

    ws1.append(['Parámetro', 'Valor'])
    for k, v in parametros_geom.items():
        ws1.append([k, v])

    fila = ws1.max_row + 2
    ws1.cell(row=fila, column=1, value="Coef. de Compacidad (fórmula)")
    ws1.cell(row=fila, column=2, value="=B6/(2*RAIZ(PI()*B2))")
    ws1.cell(row=fila+1, column=1, value="Razón de Elongación (fórmula)")
    ws1.cell(row=fila+1, column=2, value="=(2*RAIZ(B2/PI()))/B5")
    ws1.cell(row=fila+2, column=1, value="Índice de Forma (fórmula)")
    ws1.cell(row=fila+2, column=2, value="=B2/(B5^2)")

    ws2 = wb.create_sheet(title="Tiempos de Concentración")
    ws2.append(['Método', 'Tiempo (h)'])
    for k, v in tiempos_conc.items():
        ws2.append([k, v])

    chart = BarChart()
    chart.title = "Comparativa de Métodos"
    chart.x_axis.title = "Método"
    chart.y_axis.title = "Tiempo (h)"
    chart.width = 20
    chart.height = 10

    data = Reference(ws2, min_col=2, min_row=2, max_row=ws2.max_row)
    cats = Reference(ws2, min_col=1, min_row=2, max_row=ws2.max_row)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(cats)
    ws2.add_chart(chart, "E2")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
